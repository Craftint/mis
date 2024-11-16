import frappe
from frappe.utils import getdate,add_days,get_first_day,get_last_day,nowdate,flt,date_diff,add_months
import pandas as pd
import json
import os
import re

def get_setting_data(company,cost_centers,parent,parentfield,month,year_start,year_end,fiscal_year,section):
    
    set_data=[]
    cost_center_sql=''
    cost_center_str=''
    sett=frappe.db.sql("""select GROUP_CONCAT(name) as 'name',label,group_label,get_value_from,type,document,GROUP_CONCAT(accounts SEPARATOR '~') as "item_value"  
        from `tabMIS Report Account Settings` where parentfield='{0}' and parent='{1}'  group by label order by idx""".format(parentfield,parent),as_dict=1,debug=0)
    
    for se in sett:
        gpplbl=get_parent_group(se.group_label)
        item_value=str(se.item_value).split('~')
        item_value = list(map(str.strip, item_value))
        if se.type=='Account':
            acc_ar=[]            
            for ac in item_value:
                acc_ar.append(ac)
                acc_ar=acc_ar+get_child_acc(ac)
            item_value=acc_ar
        item_value_in=''
        if len(item_value):
            item_value_in='","'.join(item_value)
        
        name=str(se.name).split(',')
        name = list(map(str.strip, name))
        dp='' 
        if len(name):
            dp='","'.join(name)
        

        budsql=frappe.db.sql("""select * from `tabMIS Report Budget` b left join `tabCompany Budget` c on c.name=b.company_budget 
        where b.mis_report_account_settings in("{0}") and c.company='{1}' and c.fiscal_year='{2}' and b.label='{3}'  """.format(dp,company,fiscal_year,se.label),as_dict=1,debug=0)
        
        row={'section':section,'label':se.label,'group_label':se.group_label,'parent_group':gpplbl}
        for m in month:
            if budsql:
                mlbl=m.lower()
                bud=budsql[0].get(mlbl)                
                row.update({m:[0,bud,0]})
            else:
                row.update({m:[0,0,0]})

        if se.document=='Payroll Entry':
            
            if se.type=='Department':
                sal_p=frappe.db.sql("""select sum(gross_pay) as amt,DATE_FORMAT(posting_date, "%b") as mth from `tabSalary Slip`  
                where docstatus in (0,1) and department in("{0}") and company='{1}' and  posting_date between '{2}' and '{3}'  group by DATE_FORMAT(posting_date, "%b") order by posting_date""".format(item_value_in,company,year_start,year_end),as_dict=1)

                sal_anul=frappe.db.sql("""select sum(total_amount) as amt,DATE_FORMAT(ap.application_date, "%b") as mth from `tabAnnual Leave Payslip` ap left join `tabEmployee` e on e.name=ap.employee  
                where ap.docstatus='1' and e.department in("{0}") and e.company='{1}' and ap.application_date between '{2}' and '{3}'  group by DATE_FORMAT(ap.application_date, "%b") order by ap.application_date""".format(item_value_in,company,year_start,year_end),as_dict=1)

                
                for sal in sal_p:
                    for m in month:
                        if sal.mth==m:
                            val=[0,0,0]
                            val[0]=sal.amt
                            if budsql:
                                mlbl=m.lower()
                                val[1]=budsql[0].get(mlbl)
                            for anu in sal_anul:
                                if anu.mth==m:
                                    val[0]=float(val[0])+float(anu.amt)
                                    
                            if val[0] and val[1]:
                                val[2]=float(val[0])/float(val[1])
                        
                            row.update({m:val})

        elif se.document=='Manufacturing':
            
            msql=" "
            if se.type=='Item Group':
                msql=' and i.item_group in("'+item_value_in+'")'

            if se.type=='Item':
                msql=' and d.item_code in("'+item_value_in+'")'

            if cost_centers:
                cost_center_str='","'.join(cost_centers)

            if cost_center_str:
                cost_center_sql=' and d.cost_center in("'+cost_center_str+'") '

            stock_ent=frappe.db.sql("""select sum(amount) as amt,DATE_FORMAT(s.posting_date, "%b") as mth from `tabStock Entry` s left join `tabStock Entry Detail` d on s.name=d.parent  left join `tabItem` i on i.name=d.item_code
                where s.docstatus='1' and s.stock_entry_type='Manufacture' and s.company='{0}' and s.posting_date between '{1}' and '{2}' {3} {4} group by DATE_FORMAT(s.posting_date, "%b") order by s.posting_date""".format(company,year_start,year_end,msql,cost_center_sql),as_dict=1)
            
            for sal in stock_ent:
                for m in month:
                    if sal.mth==m:
                        val=[0,0,0]
                        val[0]=sal.amt
                        if budsql:
                            mlbl=m.lower()
                            val[1]=budsql[0].get(mlbl)
                        if val[0] and val[1]:
                            val[2]=float(val[0])/float(val[1])
                        row.update({m:val})

        elif se.document=='Stock Transfer':
            msql=" "
            if se.type=='Item Group':
                msql=' and i.item_group in("'+item_value_in+'")'

            if se.type=='Item':
                msql=' and d.item_code in("'+item_value_in+'")'

            if cost_centers:
                cost_center_str='","'.join(cost_centers)

            if cost_center_str:
                cost_center_sql=' and d.cost_center in("'+cost_center_str+'") '

            stock_ent=frappe.db.sql("""select sum(amount) as amt,DATE_FORMAT(s.posting_date, "%b") as mth from `tabStock Entry` s left join `tabStock Entry Detail` d on s.name=d.parent  left join `tabItem` i on i.name=d.item_code
                where s.docstatus='1' and s.stock_entry_type='Material Transfer' and s.company='{0}' and s.posting_date between '{1}' and '{2}' {3} {4} group by DATE_FORMAT(s.posting_date, "%b") order by s.posting_date """.format(company,year_start,year_end,msql,cost_center_sql),as_dict=1)

            
            for sal in stock_ent:
                for m in month:
                    if sal.mth==m:
                        val=[0,0,0]
                        val[0]=sal.amt
                        if budsql:                                
                            mlbl=m.lower()
                            val[1]=budsql[0].get(mlbl)
                        if val[0] and val[1]:
                            val[2]=float(val[0])/float(val[1])
                        row.update({m:val})

        else:
            if se.type=='Account':
                cost_center_sql=''
                cost_center_str=''

                if cost_centers:
                    cost_center_str='","'.join(cost_centers)

                if cost_center_str:
                    cost_center_sql=' and cost_center in("'+cost_center_str+'") '
                
                if se.get_value_from=='Debit':
                    gl_entry=frappe.db.sql("""select sum(debit)-sum(credit) as amt,DATE_FORMAT(posting_date, "%b") as mth from `tabGL Entry`  
                    where docstatus='1' and account in("{0}") and company='{1}' and posting_date between '{2}' and '{3}' {4} group by DATE_FORMAT(posting_date, "%b") order by posting_date""".format(item_value_in,company,year_start,year_end,cost_center_sql),as_dict=1)
                else:
                    gl_entry=frappe.db.sql("""select sum(credit)-sum(debit) as amt,DATE_FORMAT(posting_date, "%b") as mth from `tabGL Entry`  
                    where docstatus='1' and account in("{0}") and company='{1}' and posting_date between '{2}' and '{3}' {4} group by DATE_FORMAT(posting_date, "%b") order by posting_date """.format(item_value_in,company,year_start,year_end,cost_center_sql),as_dict=1)
                
                for sal in gl_entry:
                    for m in month:
                        if sal.mth==m:
                            val=[0,0,0]
                            val[0]=sal.amt
                            if budsql:
                                mlbl=m.lower()
                                val[1]=budsql[0].get(mlbl)
                            if val[0] and val[1]:
                                val[2]=float(val[0])/float(val[1])
                            row.update({m:val})           
        
        set_data.append(row)
            
    return set_data

@frappe.whitelist()
def get_report(company,fiscal_year,fiscal_month):
    fc=frappe.db.get_value('Fiscal Year',fiscal_year,['name','year_start_date','year_end_date'],as_dict=1)
    year_start=fc.year_start_date
    year_end=fc.year_end_date
    month=[]
    monthlist={'Jan':'JANUARY','Feb':'FEBRUARY','Mar':'MARCH','Apr':'APRIL','May':'MAY','Jun':'JUNE','Jul':'JULY','Aug':'AUGUST','Sep':'SEPTMBER','Oct':'OCTOBER','Nov':'NOVEMBER','Dec':'DECEMBER'}
    st=add_months(year_start,-1)
    while getdate(st) < get_first_day(getdate(year_end)):
        st=add_months(st, 1)
        if getdate(st)>getdate(nowdate()):
            break
        month.append(getdate(st).strftime("%b"))
        
    
    tabs=[]
    acclist=frappe.db.get_all("MIS Report Settings",filters={'company':company},fields=['name','title','page_title'],order_by='display_order')
    for ls in acclist:
        data=[]
        
        cost_centers=frappe.db.get_all("Budget Cost Center",filters={'parent':ls.name,'parenttype':'MIS Report Settings'},fields=['cost_center'],pluck='cost_center')
        
        # revenue --------------------------------------
        
        d_r=get_setting_data(company,cost_centers,ls.name,'accounts',month,year_start,year_end,fiscal_year,'revenue')
        
        # direct exp -------------------------------------
        d_d=get_setting_data(company,cost_centers,ls.name,'direct_expense',month,year_start,year_end,fiscal_year,'direct')

        # indirect direct exp -------------------------------------
        d_i=get_setting_data(company,cost_centers,ls.name,'indirect_expense',month,year_start,year_end,fiscal_year,'indirect')

        # other income -------------------------------------
        o_i=get_setting_data(company,cost_centers,ls.name,'other_income',month,year_start,year_end,fiscal_year,'otherincome')

        # other expense -------------------------------------
        o_e=get_setting_data(company,cost_centers,ls.name,'other_expense',month,year_start,year_end,fiscal_year,'otherexpense')

        # depreciation -------------------------------------
        dep=get_setting_data(company,cost_centers,ls.name,'depreciation',month,year_start,year_end,fiscal_year,'depreciation')

        #-----------------------------------------------
        #---------- row total ----------------------
        for tt in d_r:
            rtcct=0
            rtbdt=0
            rowper=0            
            for m in month:
                val=tt.get(str(m))
                rtcct=float(rtcct)+float(val[0])
                rtbdt=float(rtbdt)+float(val[1])
            if rtcct and rtbdt:
                rowper=float(rtcct)/float(rtbdt)
            tt.update({'total':[rtcct,rtbdt,rowper]})
        
        for tt in d_d:
            rtcct=0
            rtbdt=0
            rowper=0
            for m in month:
                val=tt.get(m)
                rtcct=float(rtcct)+float(val[0])
                rtbdt=float(rtbdt)+float(val[1])
            if rtcct and rtbdt:
                rowper=float(rtcct)/float(rtbdt)
            tt.update({'total':[rtcct,rtbdt,rowper]})

        for tt in d_i:
            rtcct=0
            rtbdt=0
            rowper=0
            for m in month:
                val=tt.get(m)
                rtcct=float(rtcct)+float(val[0])
                rtbdt=float(rtbdt)+float(val[1])
            if rtcct and rtbdt:
                rowper=float(rtcct)/float(rtbdt)
            tt.update({'total':[rtcct,rtbdt,rowper]})

        for tt in o_i:
            rtcct=0
            rtbdt=0
            rowper=0
            for m in month:
                val=tt.get(m)
                rtcct=float(rtcct)+float(val[0])
                rtbdt=float(rtbdt)+float(val[1])
            if rtcct and rtbdt:
                rowper=float(rtcct)/float(rtbdt)
            tt.update({'total':[rtcct,rtbdt,rowper]})

        for tt in o_e:
            rtcct=0
            rtbdt=0
            rowper=0
            for m in month:
                val=tt.get(m)
                rtcct=float(rtcct)+float(val[0])
                rtbdt=float(rtbdt)+float(val[1])
            if rtcct and rtbdt:
                rowper=float(rtcct)/float(rtbdt)
            tt.update({'total':[rtcct,rtbdt,rowper]})

        for tt in dep:
            rtcct=0
            rtbdt=0
            rowper=0
            for m in month:
                val=tt.get(m)
                rtcct=float(rtcct)+float(val[0])
                rtbdt=float(rtbdt)+float(val[1])
            if rtcct and rtbdt:
                rowper=float(rtcct)/float(rtbdt)
            tt.update({'total':[rtcct,rtbdt,rowper]})
        #---------------------------------------------------
        sectot={}
        gptot={}
        #-- colum total ----------
        for m in month:
            sectot.update({m:[0,0,0]})
            gptot.update({m:[0,0,0]})
        sectot.update({'total':[0,0,0]})
        gptot.update({'total':[0,0,0]})

        group_label=''
        d_r_t=[]
        for tt in d_r:
            
            if group_label!='' and tt.get('group_label')!=group_label:
                gplbl='Total '+str(group_label)
                row={'section':'revenue','label':gplbl,'group_label':group_label,'parent_group':'gptot'}
                
                for m in month:
                    vm=gptot.get(m)
                    if vm[0] and vm[1]:
                        tp=float(vm[0])/float(vm[1])
                        gptot.update({m:[vm[0],vm[1],tp]})
                    
                vv=gptot.get('total')
                if vv[0] and vv[1]:
                    tp=float(vv[0])/float(vv[1])
                    gptot.update({'total':[vv[0],vv[1],tp]})
                row.update(gptot)
                d_r_t.append(row)
                for m in month:
                    gptot.update({m:[0,0,0]})
                gptot.update({'total':[0,0,0]})

            for m in month:
                val=tt.get(m)
                ts=sectot.get(m)
                ts[0]=float(ts[0])+float(val[0])
                ts[1]=float(ts[1])+float(val[1])
                sectot.update({m:ts})

                tg=gptot.get(m)
                tg[0]=float(tg[0])+float(val[0])
                tg[1]=float(tg[1])+float(val[1])
                gptot.update({m:tg})

            val=tt.get('total')
            ts=sectot.get('total')
            ts[0]=float(ts[0])+float(val[0])
            ts[1]=float(ts[1])+float(val[1])
            sectot.update({'total':ts})

            tg=gptot.get('total')
            tg[0]=float(tg[0])+float(val[0])
            tg[1]=float(tg[1])+float(val[1])
            gptot.update({'total':tg})
            d_r_t.append(tt)
            group_label=tt.get('group_label')

        row={'section':'revenue','label':'Total '+str(group_label),'group_label':group_label,'parent_group':'gptot'}
        for m in month:
            vm=gptot.get(m)
            if vm[0] and vm[1]:
                tp=float(vm[0])/float(vm[1])
                gptot.update({m:[vm[0],vm[1],tp]})
                    
        vv=gptot.get('total')
        if vv[0] and vv[1]:
            tp=float(vv[0])/float(vv[1])
            gptot.update({'total':[vv[0],vv[1],tp]})
        row.update(gptot)
        d_r_t.append(row)

        row={'section':'revenue','label':'Total Revenue','group_label':'','parent_group':'setot'}
        
        for m in month:
            vm=sectot.get(m)
            if vm[0] and vm[1]:
                tp=float(vm[0])/float(vm[1])
                sectot.update({m:[vm[0],vm[1],tp]})

        vv=sectot.get('total')
        if vv[0] and vv[1]:
            tp=float(vv[0])/float(vv[1])
            sectot.update({'total':[vv[0],vv[1],tp]})
        row.update(sectot)
        d_r_t.append(row)

        #---------------------------------------------------------------
        for m in month:
            sectot.update({m:[0,0,0]})
            gptot.update({m:[0,0,0]})
        sectot.update({'total':[0,0,0]})
        gptot.update({'total':[0,0,0]})

        group_label=''
        
        d_d_t=[]
        for tt in d_d:
            
            if group_label!='' and tt.get('group_label')!=group_label:
                gplbl='Total '+str(group_label)
                row={'section':'direct','label':gplbl,'group_label':group_label,'parent_group':'gptot'}
                
                for m in month:
                    vm=gptot.get(m)
                    if vm[0] and vm[1]:
                        tp=float(vm[0])/float(vm[1])
                        gptot.update({m:[vm[0],vm[1],tp]})
                    
                vv=gptot.get('total')
                if vv[0] and vv[1]:
                    tp=float(vv[0])/float(vv[1])
                    gptot.update({'total':[vv[0],vv[1],tp]})
                row.update(gptot)
                d_d_t.append(row)
                for m in month:
                    gptot.update({m:[0,0,0]})
                gptot.update({'total':[0,0,0]})

            for m in month:
                val=tt.get(m)
                ts=sectot.get(m)
                ts[0]=float(ts[0])+float(val[0])
                ts[1]=float(ts[1])+float(val[1])
                sectot.update({m:ts})

                tg=gptot.get(m)
                tg[0]=float(tg[0])+float(val[0])
                tg[1]=float(tg[1])+float(val[1])
                gptot.update({m:tg})

            val=tt.get('total')
            ts=sectot.get('total')
            ts[0]=float(ts[0])+float(val[0])
            ts[1]=float(ts[1])+float(val[1])
            sectot.update({'total':ts})

            tg=gptot.get('total')
            tg[0]=float(tg[0])+float(val[0])
            tg[1]=float(tg[1])+float(val[1])
            gptot.update({'total':tg})
            d_d_t.append(tt)
            group_label=tt.get('group_label')

        row={'section':'direct','label':'Total '+str(group_label),'group_label':group_label,'parent_group':'gptot'}
        for m in month:
            vm=gptot.get(m)
            if vm[0] and vm[1]:
                tp=float(vm[0])/float(vm[1])
                gptot.update({m:[vm[0],vm[1],tp]})
                    
        vv=gptot.get('total')
        if vv[0] and vv[1]:
            tp=float(vv[0])/float(vv[1])
            gptot.update({'total':[vv[0],vv[1],tp]})
        row.update(gptot)
        d_d_t.append(row)

        row={'section':'direct','label':'Total Direct Expense','group_label':'','parent_group':'setot'}
        for m in month:
            vm=sectot.get(m)
            if vm[0] and vm[1]:
                tp=float(vm[0])/float(vm[1])
                sectot.update({m:[vm[0],vm[1],tp]})

        vv=sectot.get('total')
        if vv[0] and vv[1]:
            tp=float(vv[0])/float(vv[1])
            sectot.update({'total':[vv[0],vv[1],tp]})
        row.update(sectot)
        d_d_t.append(row)
        #---------------------------------------------------------------------------
        for m in month:
            sectot.update({m:[0,0,0]})
            gptot.update({m:[0,0,0]})
        sectot.update({'total':[0,0,0]})
        gptot.update({'total':[0,0,0]})
        
        group_label=''
        
        d_i_t=[]
        for tt in d_i:
            
            if group_label!='' and tt.get('group_label')!=group_label:
                gplbl='Total '+str(group_label)
                row={'section':'indirect','label':gplbl,'group_label':group_label,'parent_group':'gptot'}
                
                for m in month:
                    vm=gptot.get(m)
                    if vm[0] and vm[1]:
                        tp=float(vm[0])/float(vm[1])
                        gptot.update({m:[vm[0],vm[1],tp]})
                    
                vv=gptot.get('total')
                if vv[0] and vv[1]:
                    tp=float(vv[0])/float(vv[1])
                    gptot.update({'total':[vv[0],vv[1],tp]})
                row.update(gptot)
                d_i_t.append(row)
                for m in month:
                    gptot.update({m:[0,0,0]})
                gptot.update({'total':[0,0,0]})

            for m in month:
                val=tt.get(m)
                ts=sectot.get(m)
                ts[0]=float(ts[0])+float(val[0])
                ts[1]=float(ts[1])+float(val[1])
                sectot.update({m:ts})

                tg=gptot.get(m)
                tg[0]=float(tg[0])+float(val[0])
                tg[1]=float(tg[1])+float(val[1])
                gptot.update({m:tg})

            val=tt.get('total')
            ts=sectot.get('total')
            ts[0]=float(ts[0])+float(val[0])
            ts[1]=float(ts[1])+float(val[1])
            sectot.update({'total':ts})

            tg=gptot.get('total')
            tg[0]=float(tg[0])+float(val[0])
            tg[1]=float(tg[1])+float(val[1])
            gptot.update({'total':tg})
            d_i_t.append(tt)
            group_label=tt.get('group_label')

        row={'section':'indirect','label':'Total '+str(group_label),'group_label':group_label,'parent_group':'gptot'}
        for m in month:
            vm=gptot.get(m)
            if vm[0] and vm[1]:
                tp=float(vm[0])/float(vm[1])
                gptot.update({m:[vm[0],vm[1],tp]})
                    
        vv=gptot.get('total')
        if vv[0] and vv[1]:
            tp=float(vv[0])/float(vv[1])
            gptot.update({'total':[vv[0],vv[1],tp]})
        row.update(gptot)
        d_i_t.append(row)
        
        row={'section':'indirect','label':'Total Indirect Expense','group_label':'','parent_group':'setot'}
        for m in month:
            vm=sectot.get(m)
            if vm[0] and vm[1]:
                tp=float(vm[0])/float(vm[1])
                sectot.update({m:[vm[0],vm[1],tp]})

        vv=sectot.get('total')
        if vv[0] and vv[1]:
            tp=float(vv[0])/float(vv[1])
            sectot.update({'total':[vv[0],vv[1],tp]})
        row.update(sectot)
        d_i_t.append(row)
        #---------------------------------------------------------------------------
        for m in month:
            sectot.update({m:[0,0,0]})
            gptot.update({m:[0,0,0]})
        sectot.update({'total':[0,0,0]})
        gptot.update({'total':[0,0,0]})
        
        group_label=''
        
        o_i_t=[]
        for tt in o_i:
            
            if group_label!='' and tt.get('group_label')!=group_label:
                gplbl='Total '+str(group_label)
                row={'section':'otherincome','label':gplbl,'group_label':group_label,'parent_group':'gptot'}
                
                for m in month:
                    vm=gptot.get(m)
                    if vm[0] and vm[1]:
                        tp=float(vm[0])/float(vm[1])
                        gptot.update({m:[vm[0],vm[1],tp]})
                    
                vv=gptot.get('total')
                if vv[0] and vv[1]:
                    tp=float(vv[0])/float(vv[1])
                    gptot.update({'total':[vv[0],vv[1],tp]})
                row.update(gptot)
                o_i_t.append(row)
                for m in month:
                    gptot.update({m:[0,0,0]})
                gptot.update({'total':[0,0,0]})

            for m in month:
                val=tt.get(m)
                ts=sectot.get(m)
                ts[0]=float(ts[0])+float(val[0])
                ts[1]=float(ts[1])+float(val[1])
                sectot.update({m:ts})

                tg=gptot.get(m)
                tg[0]=float(tg[0])+float(val[0])
                tg[1]=float(tg[1])+float(val[1])
                gptot.update({m:tg})

            val=tt.get('total')
            ts=sectot.get('total')
            ts[0]=float(ts[0])+float(val[0])
            ts[1]=float(ts[1])+float(val[1])
            sectot.update({'total':ts})

            tg=gptot.get('total')
            tg[0]=float(tg[0])+float(val[0])
            tg[1]=float(tg[1])+float(val[1])
            gptot.update({'total':tg})
            o_i_t.append(tt)
            group_label=tt.get('group_label')

        row={'section':'otherincome','label':'Total '+str(group_label),'group_label':group_label,'parent_group':'gptot'}
        for m in month:
            vm=gptot.get(m)
            if vm[0] and vm[1]:
                tp=float(vm[0])/float(vm[1])
                gptot.update({m:[vm[0],vm[1],tp]})
                    
        vv=gptot.get('total')
        if vv[0] and vv[1]:
            tp=float(vv[0])/float(vv[1])
            gptot.update({'total':[vv[0],vv[1],tp]})
        row.update(gptot)
        o_i_t.append(row)
        
        row={'section':'otherincome','label':'Total Other Income','group_label':'','parent_group':'setot'}
        for m in month:
            vm=sectot.get(m)
            if vm[0] and vm[1]:
                tp=float(vm[0])/float(vm[1])
                sectot.update({m:[vm[0],vm[1],tp]})

        vv=sectot.get('total')
        if vv[0] and vv[1]:
            tp=float(vv[0])/float(vv[1])
            sectot.update({'total':[vv[0],vv[1],tp]})
        row.update(sectot)
        o_i_t.append(row)
        #---------------------------------------------------------------------------
        for m in month:
            sectot.update({m:[0,0,0]})
            gptot.update({m:[0,0,0]})
        sectot.update({'total':[0,0,0]})
        gptot.update({'total':[0,0,0]})
        
        group_label=''
        
        o_e_t=[]
        for tt in o_e:
            
            if group_label!='' and tt.get('group_label')!=group_label:
                gplbl='Total '+str(group_label)
                row={'section':'otherexpense','label':gplbl,'group_label':group_label,'parent_group':'gptot'}
                
                for m in month:
                    vm=gptot.get(m)
                    if vm[0] and vm[1]:
                        tp=float(vm[0])/float(vm[1])
                        gptot.update({m:[vm[0],vm[1],tp]})
                    
                vv=gptot.get('total')
                if vv[0] and vv[1]:
                    tp=float(vv[0])/float(vv[1])
                    gptot.update({'total':[vv[0],vv[1],tp]})
                row.update(gptot)
                o_e_t.append(row)
                for m in month:
                    gptot.update({m:[0,0,0]})
                gptot.update({'total':[0,0,0]})

            for m in month:
                val=tt.get(m)
                ts=sectot.get(m)
                ts[0]=float(ts[0])+float(val[0])
                ts[1]=float(ts[1])+float(val[1])
                sectot.update({m:ts})

                tg=gptot.get(m)
                tg[0]=float(tg[0])+float(val[0])
                tg[1]=float(tg[1])+float(val[1])
                gptot.update({m:tg})

            val=tt.get('total')
            ts=sectot.get('total')
            ts[0]=float(ts[0])+float(val[0])
            ts[1]=float(ts[1])+float(val[1])
            sectot.update({'total':ts})

            tg=gptot.get('total')
            tg[0]=float(tg[0])+float(val[0])
            tg[1]=float(tg[1])+float(val[1])
            gptot.update({'total':tg})
            o_e_t.append(tt)
            group_label=tt.get('group_label')

        row={'section':'otherexpense','label':'Total '+str(group_label),'group_label':group_label,'parent_group':'gptot'}
        for m in month:
            vm=gptot.get(m)
            if vm[0] and vm[1]:
                tp=float(vm[0])/float(vm[1])
                gptot.update({m:[vm[0],vm[1],tp]})
                    
        vv=gptot.get('total')
        if vv[0] and vv[1]:
            tp=float(vv[0])/float(vv[1])
            gptot.update({'total':[vv[0],vv[1],tp]})
        row.update(gptot)
        o_e_t.append(row)
        
        row={'section':'otherexpense','label':'Total Other Expense','group_label':'','parent_group':'setot'}
        for m in month:
            vm=sectot.get(m)
            if vm[0] and vm[1]:
                tp=float(vm[0])/float(vm[1])
                sectot.update({m:[vm[0],vm[1],tp]})

        vv=sectot.get('total')
        if vv[0] and vv[1]:
            tp=float(vv[0])/float(vv[1])
            sectot.update({'total':[vv[0],vv[1],tp]})
        row.update(sectot)
        o_e_t.append(row)
        #---------------------------------------------------------------------------
        for m in month:
            sectot.update({m:[0,0,0]})
            gptot.update({m:[0,0,0]})
        sectot.update({'total':[0,0,0]})
        gptot.update({'total':[0,0,0]})
        
        group_label=''
        
        dep_t=[]
        for tt in dep:
            
            if group_label!='' and tt.get('group_label')!=group_label:
                gplbl='Total '+str(group_label)
                row={'section':'depreciation','label':gplbl,'group_label':group_label,'parent_group':'gptot'}
                
                for m in month:
                    vm=gptot.get(m)
                    if vm[0] and vm[1]:
                        tp=float(vm[0])/float(vm[1])
                        gptot.update({m:[vm[0],vm[1],tp]})
                    
                vv=gptot.get('total')
                if vv[0] and vv[1]:
                    tp=float(vv[0])/float(vv[1])
                    gptot.update({'total':[vv[0],vv[1],tp]})
                row.update(gptot)
                dep_t.append(row)
                for m in month:
                    gptot.update({m:[0,0,0]})
                gptot.update({'total':[0,0,0]})

            for m in month:
                val=tt.get(m)
                ts=sectot.get(m)
                ts[0]=float(ts[0])+float(val[0])
                ts[1]=float(ts[1])+float(val[1])
                sectot.update({m:ts})

                tg=gptot.get(m)
                tg[0]=float(tg[0])+float(val[0])
                tg[1]=float(tg[1])+float(val[1])
                gptot.update({m:tg})

            val=tt.get('total')
            ts=sectot.get('total')
            ts[0]=float(ts[0])+float(val[0])
            ts[1]=float(ts[1])+float(val[1])
            sectot.update({'total':ts})

            tg=gptot.get('total')
            tg[0]=float(tg[0])+float(val[0])
            tg[1]=float(tg[1])+float(val[1])
            gptot.update({'total':tg})
            dep_t.append(tt)
            group_label=tt.get('group_label')

        row={'section':'depreciation','label':'Total '+str(group_label),'group_label':group_label,'parent_group':'gptot'}
        for m in month:
            vm=gptot.get(m)
            if vm[0] and vm[1]:
                tp=float(vm[0])/float(vm[1])
                gptot.update({m:[vm[0],vm[1],tp]})
                    
        vv=gptot.get('total')
        if vv[0] and vv[1]:
            tp=float(vv[0])/float(vv[1])
            gptot.update({'total':[vv[0],vv[1],tp]})
        row.update(gptot)
        dep_t.append(row)
        
        row={'section':'depreciation','label':'Total Depreciation','group_label':'','parent_group':'setot'}
        for m in month:
            vm=sectot.get(m)
            if vm[0] and vm[1]:
                tp=float(vm[0])/float(vm[1])
                sectot.update({m:[vm[0],vm[1],tp]})

        vv=sectot.get('total')
        if vv[0] and vv[1]:
            tp=float(vv[0])/float(vv[1])
            sectot.update({'total':[vv[0],vv[1],tp]})
        row.update(sectot)
        dep_t.append(row)
        #-----------------------------------------------------------------------------------
        data=d_r_t+d_d_t+d_i_t+o_i_t+o_e_t+dep_t
        tab={'title':ls.title,'page_title':ls.page_title,'data':data,'month':month}
        tabs.append(tab)
    #---------------------- consolidate ----------------------
    tabdata=[]
    constab={'title':'CONSOLIDATE','data':'','month':''}
    consdata=[]
    titles=[]
    group_labelar={'revenue':[],'direct':[],'indirect':[],'otherincome':[],'otherexpense':[],'depreciation':[]}
    for tbs in tabs:
        titles.append(tbs.get('title'))
        tbdata=[]
        for d in tbs.get('data'):
            if d.get('parent_group')=='gptot':                
                row={'section':d.get('section'),'label':d.get('group_label'),'group_label':d.get('group_label'),fiscal_month:d.get(fiscal_month),'parent_group':tbs.get('title')}
                
                sec=d.get('section')
                if d.get('group_label') and d.get('group_label') not in group_labelar.get(sec):
                    group_labelar.get(sec).append(d.get('group_label'))  
                
                tbdata.append(row)
        consdata.append(tbdata)
    
    condata=[]
    
    #frappe.msgprint(str(consdata))
    for key,cgps in group_labelar.items():
        for gps in cgps:
            row={'section':key,'label':gps,'group_label':gps}
            tbc=1
            for cc in consdata:                
                tbss='budget'+str(tbc)
                row.update({tbss:[0,0,0]})
                for c in cc:
                    if c.get('section')==key and c.get('group_label')==gps:
                        t=c.get(fiscal_month) 
                        row.update({tbss:t})
                tbc+=1
            
            condata.append(row)
       
    tbcnt=len(titles)+1    
    cocc=1
    #frappe.msgprint(str('============================'))
    #frappe.msgprint(str(condata))
    for cc in condata:
        v=[0,0,0]
        i=1
        for i in range(i,tbcnt):
            tbss='budget'+str(i)
            conval=cc.get(tbss)
            v[0]=float(v[0])+float(conval[0])
            v[1]=float(v[1])+float(conval[1])
            
        if v[0] and v[1]:
            v[2]=v[0]/v[1]
        cc.update({'total':v})
    #frappe.msgprint(str('============================'))
    #frappe.msgprint(str(condata))
    #frappe.msgprint(str('============================'))
    
    revtot={'section':'revenue','label':'Total Revenue','group_label':'','parent_group':''}
    i=1
    for i in range(i,tbcnt):
        tbss='budget'+str(i)
        revtot.update({tbss:[0,0,0]})
    revtot.update({'total':[0,0,0]})
    
    
    dirtot={'section':'direct','label':'Total Direct Expense','group_label':'','parent_group':''}
    i=1
    for i in range(i,tbcnt):
        tbss='budget'+str(i)
        dirtot.update({tbss:[0,0,0]})
    dirtot.update({'total':[0,0,0]})

    
    indirtot={'section':'indirect','label':'Total indirect Expense','group_label':'','parent_group':''}
    i=1
    for i in range(i,tbcnt):
        tbss='budget'+str(i)
        indirtot.update({tbss:[0,0,0]})
    indirtot.update({'total':[0,0,0]})

    
    othinctot={'section':'otherincome','label':'Total Other Income','group_label':'','parent_group':''}
    i=1
    for i in range(i,tbcnt):
        tbss='budget'+str(i)
        othinctot.update({tbss:[0,0,0]})
    othinctot.update({'total':[0,0,0]})

    othexptot={'section':'otherexpense','label':'Total Other Expense','group_label':'','parent_group':''}
    i=1
    for i in range(i,tbcnt):
        tbss='budget'+str(i)
        othexptot.update({tbss:[0,0,0]})
    othexptot.update({'total':[0,0,0]})

    depetot={'section':'depreciation','label':'Total Depreciation','group_label':'','parent_group':''}
    i=1
    for i in range(i,tbcnt):
        tbss='budget'+str(i)
        depetot.update({tbss:[0,0,0]})
    depetot.update({'total':[0,0,0]})

    revar=[]
    for cc in condata:
        if cc.get('section')=='revenue':
            revar.append(cc)
    

    dirar=[]
    for cc in condata:
        if cc.get('section')=='direct':
            dirar.append(cc)
    

    indirar=[]
    for cc in condata:
        if cc.get('section')=='indirect':
            indirar.append(cc)

    othincrar=[]
    for cc in condata:
        if cc.get('section')=='otherincome':
            othincrar.append(cc)
    
    othexprar=[]
    for cc in condata:
        if cc.get('section')=='otherexpense':
            othexprar.append(cc)

    deprar=[]
    for cc in condata:
        if cc.get('section')=='depreciation':
            deprar.append(cc)

    #frappe.msgprint(str(indirar))
    #frappe.msgprint(str('============================'))
    
    i=1
    for i in range(i,tbcnt):
        tbss='budget'+str(i)
        
        v1=sum(p.get(tbss)[0] for p in revar)
        v2=sum(p.get(tbss)[1] for p in revar)

        v3=sum(p.get(tbss)[0] for p in dirar)
        v4=sum(p.get(tbss)[1] for p in dirar)

        v5=sum(p.get(tbss)[0] for p in indirar)
        v6=sum(p.get(tbss)[1] for p in indirar)

        v7=sum(p.get(tbss)[0] for p in othincrar)
        v8=sum(p.get(tbss)[1] for p in othincrar)

        v9=sum(p.get(tbss)[0] for p in othexprar)
        v10=sum(p.get(tbss)[1] for p in othexprar)

        v11=sum(p.get(tbss)[0] for p in deprar)
        v12=sum(p.get(tbss)[1] for p in deprar)

        vv=0
        if v1 and v2:
            vv=v1/v2
        revtot.update({tbss:[v1,v2,vv]})

        vv=0
        if v3 and v4:
            vv=v3/v4
        dirtot.update({tbss:[v3,v4,vv]})

        vv=0
        if v5 and v6:
            vv=v5/v6
        indirtot.update({tbss:[v5,v6,vv]})

        vv=0
        if v7 and v8:
            vv=v7/v8
        othinctot.update({tbss:[v7,v8,vv]})

        vv=0
        if v9 and v10:
            vv=v9/v10
        othexptot.update({tbss:[v9,v10,vv]})

        vv=0
        if v11 and v12:
            vv=v11/v12
        depetot.update({tbss:[v11,v12,vv]})

    v1=sum(p.get('total')[0] for p in revar)
    v2=sum(p.get('total')[1] for p in revar)

    v3=sum(p.get('total')[0] for p in dirar)
    v4=sum(p.get('total')[1] for p in dirar)

    v5=sum(p.get('total')[0] for p in indirar)
    v6=sum(p.get('total')[1] for p in indirar)
    
    v7=sum(p.get('total')[0] for p in othincrar)
    v8=sum(p.get('total')[1] for p in othincrar)

    v9=sum(p.get('total')[0] for p in othexprar)
    v10=sum(p.get('total')[1] for p in othexprar)

    v11=sum(p.get('total')[0] for p in deprar)
    v12=sum(p.get('total')[1] for p in deprar)

    vv=0
    if v1 and v2:
        vv=v1/v2
    revtot.update({'total':[v1,v2,vv]})

    vv=0
    if v3 and v4:
        vv=v3/v4
    dirtot.update({'total':[v3,v4,vv]})

    vv=0
    if v5 and v6:
        vv=v5/v6
    indirtot.update({'total':[v5,v6,vv]})

    vv=0
    if v7 and v8:
        vv=v7/v8
    othinctot.update({'total':[v7,v8,vv]})

    vv=0
    if v9 and v10:
        vv=v9/v10
    othexptot.update({'total':[v9,v10,vv]})

    vv=0
    if v11 and v12:
        vv=v11/v12
    depetot.update({'total':[v11,v12,vv]})


    revar.append(revtot)
    dirar.append(dirtot)
    indirar.append(indirtot)
    othincrar.append(othinctot)
    othexprar.append(othexptot)
    deprar.append(depetot)
    #frappe.msgprint(str(deprar))
    gprofit={'section':'','label':'GROSS PROFIT','group_label':'','parent_group':''}
    oprofit={'section':'','label':'NET OPERATING PROFIT / (LOSS) BEFORE DEPRECIATION & INTEREST','group_label':'','parent_group':''}
    oinprofit={'section':'','label':'PROFIT BEFORE DEPR & INTRESEST','group_label':'','parent_group':''}
    oexprofit={'section':'','label':'NET OPERATING PROFIT / (LOSS) EXCLUDING DEPRECIATION','group_label':'','parent_group':''}
    nprofit={'section':'','label':'PROFIT/LOSS','group_label':'','parent_group':''}
    
    i=1
    for i in range(i,tbcnt):
        tbss='budget'+str(i)
        rv=revtot.get(tbss)
        dr=dirtot.get(tbss)
        idr=indirtot.get(tbss)
        othi=othinctot.get(tbss)
        othe=othexptot.get(tbss)
        dp=depetot.get(tbss)

        c=0
        a=rv[0]-dr[0]
        b=rv[1]-dr[1]
        if b and a:
            c=a/b
        gprofit.update({tbss:[a,b,c]})
        #-------------------------------------
        a=a-idr[0]
        b=b-idr[1]
        c=0
        if b and a:
            c=a/b
        oprofit.update({tbss:[a,b,c]})
        #-------------------------------------
        a=a+othi[0]
        b=b+othi[1]
        c=0
        if b and a:
            c=a/b
        oinprofit.update({tbss:[a,b,c]})
        #-------------------------------------
        a=a-othe[0]
        b=b-othe[1]
        c=0
        if b and a:
            c=a/b
        oexprofit.update({tbss:[a,b,c]})
        #-------------------------------------
        a=a-dp[0]
        b=b-dp[1]
        c=0
        if b and a:
            c=a/b
        nprofit.update({tbss:[a,b,c]})
        #-------------------------------------

    rv=revtot.get('total')
    dr=dirtot.get('total')
    idr=indirtot.get('total')
    othi=othinctot.get('total')
    othe=othexptot.get('total')
    dp=depetot.get('total')

    c=0
    a=rv[0]-dr[0]
    b=rv[1]-dr[1]
    if b and a:
        c=a/b
    gprofit.update({'total':[a,b,c]})
    #-------------------------------------
    a=a-idr[0]
    b=b-idr[1]
    c=0
    if b and a:
        c=a/b
    oprofit.update({'total':[a,b,c]})
    #-------------------------------------
    a=a+othi[0]
    b=b+othi[1]
    c=0
    if b and a:
        c=a/b
    oinprofit.update({'total':[a,b,c]})
    #-------------------------------------
    a=a-othe[0]
    b=b-othe[1]
    c=0
    if b and a:
        c=a/b
    oexprofit.update({'total':[a,b,c]})
    #-------------------------------------
    a=a-dp[0]
    b=b-dp[1]
    c=0
    if b and a:
        c=a/b
    nprofit.update({'total':[a,b,c]})
    netprofit=a

    dirar.append(gprofit)
    indirar.append(oprofit)
    othincrar.append(oinprofit)
    othexprar.append(oexprofit)
    deprar.append(nprofit)

    consoli=revar+dirar+indirar+othincrar+othexprar+deprar
    #frappe.msgprint(str(consoli))
    
    #-------------------------------------------------------------------------------
    bugtit=[]
    bugtit.append('DEPARTMENT SUMMARY FOR MONTH : '+str(monthlist.get(fiscal_month))+'-'+str(fiscal_year))
    budget0='<span style="padding:15px; display: block;"><b>DEPARTMENT SUMMARY FOR MONTH : '+str(monthlist.get(fiscal_month))+'-'+str(fiscal_year)+'</b></span>'
    budget0+= '<div style="overflow-x: auto;">'
    budget0+='<table class="table table-bordered" id="budgets"><thead> '
    budget0+='<tr class="table-secondary"><th scope="col" style="width:150px;"></th>'
    sbt='<tr class="table-secondary"><td scope="col"></td>'
    headconsol=['']
    for ti in titles:
        headconsol.append(str(ti))
        headconsol.append('')
        headconsol.append('')
        budget0+='<th scope="col" colspan=3 class="text-center">'+str(ti)+'</th>'
        sbt+='<td scope="col">ACTUAL</td><td scope="col">BUDGET</td><td scope="col">0%</td>'

    headconsol.append('Total')
    headconsol.append('')
    headconsol.append('')
    budget0+='<th scope="col" colspan=3 class="text-center">Total</th> </tr></thead> '
    sbt+='<td scope="col">ACTUAL</td><td scope="col">BUDGET</td><td scope="col">0%</td></tr>'
    budget0+=sbt
    
    for cn in consoli:
        hd='  '
        if cn.get('label') and not cn.get('section'):
            hd=' class="table-secondary"'

        if cn.get('label') and not cn.get('group_label'):
            hd=' class="table-secondary"'

        budget0+='<tr '+str(hd)+'>'
        budget0+='<td style="width:150px;">'+str(cn.get('label'))+'</td>'

        
        i=1
        blk=''
        for i in range(i,tbcnt):
            tbss='budget'+str(i)
            v=cn.get(tbss)
            if v:
                budget0+='<td class="text-right">'+str(frappe.utils.fmt_money(v[0]))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(v[1]))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(v[2]))+'</td>'
            else:
                budget0+='<td>0</td><td>0</td><td>0</td>'
            blk+='<td class="text-right"></td><td class="text-right"></td><td class="text-right"></td>'
        v=cn.get('total')
        if v:
            budget0+='<td class="text-right">'+str(frappe.utils.fmt_money(v[0]))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(v[1]))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(v[2]))+'</td>'
        else:
            budget0+='<td>0</td><td>0</td><td>0</td>'
        budget0+='</tr>'

    
    #budget0+='<tr >'
    #budget0+='<td >DEPRECIATION</td>'
    #budget0+=blk
    #expense_account=frappe.db.get_value('company',company,'depreciation_expense_account')
    #gl_entry=frappe.db.sql("""select IFNULL(sum(debit),0) as amt from `tabGL Entry`  
    #                    where docstatus='1' and account in("{0}") and company='{1}' and posting_date between '{2}' and '{3}' """.format(expense_account,company,year_start,year_end),as_dict=1)
    #if gl_entry:
    #    budget0+='<td class="text-right">'+str(gl_entry[0].amt)+'</td><td class="text-right">0</td><td class="text-right">0</td>'
    #else:
    #    budget0+='<td class="text-right">0</td><td class="text-right">0</td><td class="text-right">0</td>'
    
    #budget0+= '</tr>'

    #budget0+='<tr class="table-secondary">'
    #budget0+='<td >AFTER DEPRECIATION PROFIT /LOSS</td>'
    #budget0+=blk    
    #budget0+='<td class="text-right">'+str(frappe.utils.fmt_money(float(netprofit)-float(gl_entry[0].amt)))+'</td><td class="text-right">0</td><td class="text-right">0</td>'
    #budget0+= '</tr>'

    budget0+= '</table></div>'  
    
    

    repdatahtml={'budget0':budget0}
    dcon=[]
    
    df = pd.read_html(budget0,header=0)[0]
    
    dcon.append(bugtit)
    dcon.append(headconsol)
    dcon=dcon+df.values.tolist()
    
    for index, val in enumerate(dcon):
        for index2, valu in enumerate(val):
            
            pattern = r"^[-+]?[0-9]*\.?[0-9]+$"
            match = re.match(pattern, str(dcon[index][index2]))
            if match:
                dcon[index][index2]=float(dcon[index][index2])
                
    
    nam=str(frappe.session.user)+'CONSOLIDATE'
    consolid=json.dumps(dcon)
    cache = frappe.cache()
    cache.set(nam, consolid) 
    
    
    tbc=1
    html=''
    for tbd in tabs:
        headconsol=['']
        headconsol2=['']
        bugtit=[]
        bugtit.append(str(tbd.get('page_title'))+' FOR THE MONTH OF  '+str(monthlist.get(fiscal_month))+' - '+str(fiscal_year))
        html='<span style="padding:15px;display: block;"><b>'+str(tbd.get('page_title'))+' FOR THE MONTH OF  '+str(monthlist.get(fiscal_month))+' - '+str(fiscal_year)+'</b></span>'
        html+= '<div style="overflow-x: auto;"><table class="table table-bordered" >'
        html+='<tr class="table-secondary"><th scope="col" style="width:150px;"></th>'
        th='<tr class="table-secondary"><td scope="col"></td>'
        colcnt=len(tbd.get('month'))+2
        
        last_mc=len(tbd.get('month'))-1
        first_mth=tbd.get('month')[0]
        last_mth=tbd.get('month')[last_mc]
        prev_mth=''
        curr_mth=''
        for m in tbd.get('month'):
            if fiscal_month==m:
                curr_mth=m
                break
            prev_mth=m

        if prev_mth or curr_mth:
            
            if curr_mth:
                html+='<th scope="col" class="text-center" colspan=3> Current Month '+str(curr_mth)+'</th>'
                th+='<td scope="col">ACTUAL</td><td scope="col">BUDGET</td><td scope="col">0%</td>'
                headconsol.append('Current Month '+str(curr_mth))
                headconsol.append('')
                headconsol.append('')
            
            if prev_mth:
                html+='<th scope="col" class="text-center" colspan=3>Previous Month '+str(prev_mth)+'</th>'
                th+='<td scope="col">ACTUAL</td><td scope="col">BUDGET</td><td scope="col">0%</td>'
                headconsol.append('Previous Month '+str(prev_mth))
                headconsol.append('')
                headconsol.append('')

            html+='<th scope="col" class="text-center" colspan=3>'+str(first_mth)+' To '+str(last_mth)+'</th><td> &nbsp;&nbsp;</td>'
            th+='<td scope="col">ACTUAL</td><td scope="col">BUDGET</td><td scope="col">0%</td><td></td>'
            headconsol.append(str(first_mth)+' To '+str(last_mth))
            headconsol.append('')
            headconsol.append('')
        th2='<td></td>'
        headconsol.append('')
        html2=''
        for m in tbd.get('month'):
            html+='<th scope="col" class="text-center" >ACTUAL</th>'
            headconsol.append('ACTUAL')
            html2+='<th scope="col" class="text-center" >BUDGET</th>'
            headconsol2.append('BUDGET')
            th+='<td scope="col">'+str(m)+'</td>'
            th2+='<td scope="col">'+str(m)+'</td>'
        html+='<th scope="col" class="text-center">ACTUAL</th><th>&nbsp;</th>'
        html2+='<th scope="col" class="text-center">BUDGET</th>'
        headconsol.append('ACTUAL')
        headconsol2.append('BUDGET')
        html+=html2
        html+='</tr>'
        th+='<td scope="col">Total</td>'
        th2+='<td scope="col">Total</td></tr>'

        html+=th+th2
        group_label=''
        totals=[]
        for d in tbd.get('data'):
            head=''
            if not d.get('group_label'):
                head=' class="table-secondary"'
                

            if group_label=='' or (group_label!='' and d.get('group_label')!='' and group_label!=d.get('group_label')):
                html+='<tr ><td scope="col"><b>'+str(d.get('group_label'))+'</b></td>'
                i=0
                if prev_mth or curr_mth:
                    if curr_mth:
                        html+='<td class="text-right bglight"></td><td class="text-right bglight"></td><td class="text-right bglight"></td>'
                    if prev_mth:
                        html+='<td></td><td></td><td></td>'
                html+='<td class="text-right bgdark"></td><td class="text-right bgdark"></td><td class="text-right bgdark"></td>'

                for i in range(i,colcnt):
                    html+='<td></td><td></td>'
                
                html+='</tr>'

            if d.get('group_label'):
                group_label=d.get('group_label')

               
            

            html+='<tr '+str(head)+'>'
            html+='<td >'+str(d.get('label'))+'</td>'

            if prev_mth or curr_mth:
            
                if curr_mth:
                    v=d.get(curr_mth)
                    if v:
                        html+='<td class="text-right bglight">'+str(frappe.utils.fmt_money(v[0]))+'</td><td class="text-right bglight">'+str(frappe.utils.fmt_money(v[1]))+'</td><td class="text-right bglight">'+str(frappe.utils.fmt_money(v[2]))+'%</td>'
                    else:
                        html+='<td>0</td><td>0</td><td>0</td>'
                
                if prev_mth:
                    v=d.get(prev_mth)
                    if v:
                        html+='<td class="text-right">'+str(frappe.utils.fmt_money(v[0]))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(v[1]))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(v[2]))+'%</td>'
                    else:
                        html+='<td>0</td><td>0</td><td>0</td>'

                v=d.get('total')
                if v:
                    html+='<td class="text-right bgdark">'+str(frappe.utils.fmt_money(v[0]))+'</td><td class="text-right bgdark">'+str(frappe.utils.fmt_money(v[1]))+'</td><td class="text-right bgdark">'+str(frappe.utils.fmt_money(v[2]))+'%</td><td></td>'
                else:
                    html+='<td>0</td><td>0</td><td>0</td>'

            html2=''
            for m in tbd.get('month'):
                v=d.get(m)
                if v:
                    html+='<td class="text-right">'+str(frappe.utils.fmt_money(v[0]))+'</td>'
                    html2+='<td class="text-right">'+str(frappe.utils.fmt_money(v[1]))+'</td>'
                else:
                    html+='<td>0</td>'
                    html2+='<td>0</td>'

            v=d.get('total')
            if v:
                html+='<td class="text-right">'+str(frappe.utils.fmt_money(v[0]))+'</td>'
                html2+='<td class="text-right">'+str(frappe.utils.fmt_money(v[1]))+'</td>'
            else:
                html+='<td>0</td>'
                html2+='<td>0</td>'
            
            html+='<td></td>'+html2+'</tr>'

            if d.get('parent_group')=='setot':
                
                if len(totals)==1:
                    html2=''
                    html+='<tr class="table-secondary" ><td scope="col"><b>Gross Profit</b></td>'
                    prof=totals[0]

                    if prev_mth or curr_mth:
            
                        if curr_mth:
                            v1=prof.get(curr_mth)
                            v2=d.get(curr_mth)
                            act=v1[0]-v2[0]
                            bud=v1[1]-v2[1]
                            p=0
                            if act and bud:
                                p=float(act)/float(bud)
                            html+='<td class="text-right">'+str(frappe.utils.fmt_money(act))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(bud))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(p))+'%</td>'

                        
                        if prev_mth:
                            v1=prof.get(prev_mth)
                            v2=d.get(prev_mth)
                            act=v1[0]-v2[0]
                            bud=v1[1]-v2[1]
                            p=0
                            if act and bud:
                                p=float(act)/float(bud)
                            html+='<td class="text-right">'+str(frappe.utils.fmt_money(act))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(bud))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(p))+'%</td>'


                        v1=prof.get('total')
                        v2=d.get('total')
                        act=v1[0]-v2[0]
                        bud=v1[1]-v2[1]
                        p=0
                        if act and bud:
                            p=float(act)/float(bud)
                        html+='<td class="text-right">'+str(frappe.utils.fmt_money(act))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(bud))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(p))+'%</td><td></td>'


                    for m in tbd.get('month'):
                        v1=prof.get(m)
                        v2=d.get(m)
                        act=v1[0]-v2[0]
                        bud=v1[1]-v2[1]
                        p=0
                        if act and bud:
                            p=float(act)/float(bud)
                        html+='<td class="text-right">'+str(frappe.utils.fmt_money(act))+'</td>'
                        html2+='<td class="text-right">'+str(frappe.utils.fmt_money(bud))+'</td>'
                    v1=prof.get('total')
                    v2=d.get('total')
                    act=v1[0]-v2[0]
                    bud=v1[1]-v2[1]
                    p=0
                    if act and bud:
                        p=float(act)/float(bud)
                    html+='<td class="text-right">'+str(frappe.utils.fmt_money(act))+'</td>'
                    html2+='<td class="text-right">'+str(frappe.utils.fmt_money(bud))+'</td>'

                    html+='<td></td>'+html2+'</tr >'

                if len(totals)==2:
                    html+='<tr class="table-secondary" ><td scope="col"><b>NET OPERATING PROFIT / (LOSS) BEFORE DEPRECIATION & INTEREST </b></td>'
                    html2=''
                    prof=totals[0]
                    gprof=totals[1]

                    if prev_mth or curr_mth:
                        if curr_mth:
                            v1=prof.get(curr_mth)
                            v2=gprof.get(curr_mth)
                            v3=d.get(curr_mth)
                            act=v1[0]-v2[0]-v3[0]
                            bud=v1[1]-v2[1]-v3[1]
                            if act and bud:
                                p=float(act)/float(bud)
                            html+='<td class="text-right">'+str(frappe.utils.fmt_money(act))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(bud))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(p))+'%</td>'

                        if prev_mth:
                            v1=prof.get(prev_mth)
                            v2=gprof.get(prev_mth)
                            v3=d.get(prev_mth)
                            act=v1[0]-v2[0]-v3[0]
                            bud=v1[1]-v2[1]-v3[1]
                            if act and bud:
                                p=float(act)/float(bud)
                            html+='<td class="text-right">'+str(frappe.utils.fmt_money(act))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(bud))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(p))+'%</td>'

                        
                        v1=prof.get('total')
                        v2=gprof.get('total')
                        v3=d.get('total')
                        act=v1[0]-v2[0]-v3[0]
                        bud=v1[1]-v2[1]-v3[1]
                        if act and bud:
                            p=float(act)/float(bud)
                        html+='<td class="text-right">'+str(frappe.utils.fmt_money(act))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(bud))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(p))+'%</td><td></td>'


                    for m in tbd.get('month'):
                        v1=prof.get(m)
                        v2=gprof.get(m)
                        v3=d.get(m)
                        act=v1[0]-v2[0]-v3[0]
                        bud=v1[1]-v2[1]-v3[1]
                        if act and bud:
                            p=float(act)/float(bud)
                        html+='<td class="text-right">'+str(frappe.utils.fmt_money(act))+'</td>'
                        html2+='<td class="text-right">'+str(frappe.utils.fmt_money(bud))+'</td>'
                    v1=prof.get('total')
                    v2=gprof.get('total')
                    v3=d.get('total')
                    act=v1[0]-v2[0]-v3[0]
                    bud=v1[1]-v2[1]-v3[1]
                    if act and bud:
                        p=float(act)/float(bud)
                    html+='<td class="text-right">'+str(frappe.utils.fmt_money(act))+'</td>'
                    html2+='<td class="text-right">'+str(frappe.utils.fmt_money(bud))+'</td>'
                    html+='<td></td>'+html2+'</tr >'

                if len(totals)==3:
                    html+='<tr class="table-secondary" ><td scope="col"><b>PROFIT BEFORE DEPR & INTRESEST  </b></td>'
                    html2=''
                    prof=totals[0]
                    gprof=totals[1]
                    ind=totals[2]

                    if prev_mth or curr_mth:
                        if curr_mth:
                            v1=prof.get(curr_mth)
                            v2=gprof.get(curr_mth)
                            v3=ind.get(curr_mth)
                            v4=d.get(curr_mth)
                            act=v1[0]-v2[0]-v3[0]+v4[0]
                            bud=v1[1]-v2[1]-v3[1]+v4[1]
                            if act and bud:
                                p=float(act)/float(bud)
                            html+='<td class="text-right">'+str(frappe.utils.fmt_money(act))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(bud))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(p))+'%</td>'

                        if prev_mth:
                            v1=prof.get(prev_mth)
                            v2=gprof.get(prev_mth)
                            v3=ind.get(prev_mth)
                            v4=d.get(prev_mth)
                            act=v1[0]-v2[0]-v3[0]+v4[0]
                            bud=v1[1]-v2[1]-v3[1]+v4[1]
                            if act and bud:
                                p=float(act)/float(bud)
                            html+='<td class="text-right">'+str(frappe.utils.fmt_money(act))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(bud))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(p))+'%</td>'

                        
                        v1=prof.get('total')
                        v2=gprof.get('total')
                        v3=ind.get('total')
                        v4=d.get('total')
                        act=v1[0]-v2[0]-v3[0]+v4[0]
                        bud=v1[1]-v2[1]-v3[1]+v4[1]
                        if act and bud:
                            p=float(act)/float(bud)
                        html+='<td class="text-right">'+str(frappe.utils.fmt_money(act))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(bud))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(p))+'%</td><td></td>'


                    for m in tbd.get('month'):
                        v1=prof.get(m)
                        v2=gprof.get(m)
                        v3=ind.get(m)
                        v4=d.get(m)
                        act=v1[0]-v2[0]-v3[0]+v4[0]
                        bud=v1[1]-v2[1]-v3[1]+v4[1]
                        if act and bud:
                            p=float(act)/float(bud)
                        html+='<td class="text-right">'+str(frappe.utils.fmt_money(act))+'</td>'
                        html2+='<td class="text-right">'+str(frappe.utils.fmt_money(bud))+'</td>'
                    v1=prof.get('total')
                    v2=gprof.get('total')
                    v3=ind.get('total')
                    v4=d.get('total')
                    act=v1[0]-v2[0]-v3[0]+v4[0]
                    bud=v1[1]-v2[1]-v3[1]+v4[1]
                    if act and bud:
                        p=float(act)/float(bud)
                    html+='<td class="text-right">'+str(frappe.utils.fmt_money(act))+'</td>'
                    html2+='<td class="text-right">'+str(frappe.utils.fmt_money(bud))+'</td>'
                    html+='<td></td>'+html2+'</tr >'

                if len(totals)==4:
                    html+='<tr class="table-secondary" ><td scope="col"><b>NET OPERATING PROFIT / (LOSS) EXCLUDING  DEPRECIATION</b></td>'
                    html2=''
                    prof=totals[0]
                    gprof=totals[1]
                    ind=totals[2]
                    othinc=totals[3]

                    if prev_mth or curr_mth:
                        if curr_mth:
                            v1=prof.get(curr_mth)
                            v2=gprof.get(curr_mth)
                            v3=ind.get(curr_mth)
                            v4=othinc.get(curr_mth)
                            v5=d.get(curr_mth)
                            act=v1[0]-v2[0]-v3[0]+v4[0]-v5[0]
                            bud=v1[1]-v2[1]-v3[1]+v4[1]-v5[1]
                            if act and bud:
                                p=float(act)/float(bud)
                            html+='<td class="text-right">'+str(frappe.utils.fmt_money(act))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(bud))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(p))+'%</td>'

                        if prev_mth:
                            v1=prof.get(prev_mth)
                            v2=gprof.get(prev_mth)
                            v3=ind.get(prev_mth)
                            v4=othinc.get(prev_mth)
                            v5=d.get(prev_mth)
                            act=v1[0]-v2[0]-v3[0]+v4[0]-v5[0]
                            bud=v1[1]-v2[1]-v3[1]+v4[1]-v5[1]
                            if act and bud:
                                p=float(act)/float(bud)
                            html+='<td class="text-right">'+str(frappe.utils.fmt_money(act))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(bud))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(p))+'%</td>'

                        
                        v1=prof.get('total')
                        v2=gprof.get('total')
                        v3=ind.get('total')
                        v4=othinc.get('total')
                        v5=d.get('total')
                        act=v1[0]-v2[0]-v3[0]+v4[0]-v5[0]
                        bud=v1[1]-v2[1]-v3[1]+v4[1]-v5[1]
                        if act and bud:
                            p=float(act)/float(bud)
                        html+='<td class="text-right">'+str(frappe.utils.fmt_money(act))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(bud))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(p))+'%</td><td></td>'


                    for m in tbd.get('month'):
                        v1=prof.get(m)
                        v2=gprof.get(m)
                        v3=ind.get(m)
                        v4=othinc.get(m)
                        v5=d.get(m)
                        act=v1[0]-v2[0]-v3[0]+v4[0]-v5[0]
                        bud=v1[1]-v2[1]-v3[1]+v4[1]-v5[1]
                        if act and bud:
                            p=float(act)/float(bud)
                        html+='<td class="text-right">'+str(frappe.utils.fmt_money(act))+'</td>'
                        html2+='<td class="text-right">'+str(frappe.utils.fmt_money(bud))+'</td>'
                    v1=prof.get('total')
                    v2=gprof.get('total')
                    v3=ind.get('total')
                    v4=othinc.get('total')
                    v5=d.get('total')
                    act=v1[0]-v2[0]-v3[0]+v4[0]-v5[0]
                    bud=v1[1]-v2[1]-v3[1]+v4[1]-v5[1]
                    if act and bud:
                        p=float(act)/float(bud)
                    html+='<td class="text-right">'+str(frappe.utils.fmt_money(act))+'</td>'
                    html2+='<td class="text-right">'+str(frappe.utils.fmt_money(bud))+'</td>'
                    html+='<td></td>'+html2+'</tr >'

                if len(totals)==5:
                    
                    html+='<tr class="table-secondary" ><td scope="col"><b>Operation Profit + Other Income (FD)</b></td>' 
                    html3='<tr class="table-secondary" ><td scope="col"><b>Zakat Expense @2.75%</b></td>'
                    html4='<tr class="table-secondary" ><td scope="col"><b>NET PROFIT / (LOSS)</b></td>'
                    html2=''
                    html32=''
                    html42=''
                    prof=totals[0]
                    gprof=totals[1]
                    ind=totals[2]
                    othinc=totals[3]
                    othexp=totals[4]

                    if prev_mth or curr_mth:
                        if curr_mth:
                            v1=prof.get(curr_mth)
                            v2=gprof.get(curr_mth)
                            v3=ind.get(curr_mth)
                            v4=othinc.get(curr_mth)
                            v5=othexp.get(curr_mth)
                            v6=d.get(curr_mth)
                            act=v1[0]-v2[0]-v3[0]+v4[0]-v5[0]-v6[0]
                            bud=v1[1]-v2[1]-v3[1]+v4[1]-v5[1]-v6[1]
                            if act and bud:
                                p=float(act)/float(bud)

                            ps=0
                            psn=0
                            if act > 0 :
                                acts=float(act)*.0275
                                buds=float(bud)*.0275
                                
                                if acts and buds:
                                    ps=float(acts)/float(buds)

                                actsn=act-acts
                                budsn=bud-buds
                                
                                if actsn and budsn:
                                    psn=float(actsn)/float(budsn)
                            else:
                                acts=0
                                buds=0
                                actsn=act
                                budsn=bud

                            html+='<td class="text-right">'+str(frappe.utils.fmt_money(act))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(bud))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(p))+'%</td>'
                            html3+='<td class="text-right">'+str(frappe.utils.fmt_money(acts))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(buds))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(ps))+'%</td>'
                            html4+='<td class="text-right">'+str(frappe.utils.fmt_money(actsn))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(budsn))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(psn))+'%</td>'

                        if prev_mth:
                            v1=prof.get(prev_mth)
                            v2=gprof.get(prev_mth)
                            v3=ind.get(prev_mth)
                            v4=othinc.get(prev_mth)
                            v5=othexp.get(prev_mth)
                            v6=d.get(prev_mth)
                            act=v1[0]-v2[0]-v3[0]+v4[0]-v5[0]-v6[0]
                            bud=v1[1]-v2[1]-v3[1]+v4[1]-v5[1]-v6[1]
                            if act and bud:
                                p=float(act)/float(bud)

                            ps=0
                            psn=0
                            if act > 0 :
                                acts=float(act)*.0275
                                buds=float(bud)*.0275
                                
                                if acts and buds:
                                    ps=float(acts)/float(buds)

                                actsn=act-acts
                                budsn=bud-buds
                                
                                if actsn and budsn:
                                    psn=float(actsn)/float(budsn)
                            else:
                                acts=0
                                buds=0
                                actsn=act
                                budsn=bud

                            html+='<td class="text-right">'+str(frappe.utils.fmt_money(act))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(bud))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(p))+'%</td>'
                            html3+='<td class="text-right">'+str(frappe.utils.fmt_money(acts))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(buds))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(ps))+'%</td>'
                            html4+='<td class="text-right">'+str(frappe.utils.fmt_money(actsn))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(budsn))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(psn))+'%</td>'
                        
                        v1=prof.get('total')
                        v2=gprof.get('total')
                        v3=ind.get('total')
                        v4=othinc.get('total')
                        v5=othexp.get('total')
                        v6=d.get('total')
                        act=v1[0]-v2[0]-v3[0]+v4[0]-v5[0]-v6[0]
                        bud=v1[1]-v2[1]-v3[1]+v4[1]-v5[1]-v6[1]
                        if act and bud:
                            p=float(act)/float(bud)

                        ps=0
                        psn=0
                        if act > 0 :
                            acts=float(act)*.0275
                            buds=float(bud)*.0275
                            
                            if acts and buds:
                                ps=float(acts)/float(buds)

                            actsn=act-acts
                            budsn=bud-buds
                            
                            if actsn and budsn:
                                psn=float(actsn)/float(budsn)
                        else:
                            acts=0
                            buds=0
                            actsn=act
                            budsn=bud

                        html+='<td class="text-right">'+str(frappe.utils.fmt_money(act))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(bud))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(p))+'%</td><td></td>'
                        html3+='<td class="text-right">'+str(frappe.utils.fmt_money(acts))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(buds))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(ps))+'%</td>'
                        html4+='<td class="text-right">'+str(frappe.utils.fmt_money(actsn))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(budsn))+'</td><td class="text-right">'+str(frappe.utils.fmt_money(psn))+'%</td>'

                    for m in tbd.get('month'):
                        v1=prof.get(m)
                        v2=gprof.get(m)
                        v3=ind.get(m)
                        v4=othinc.get(m)
                        v5=othexp.get(m)
                        v6=d.get(m)
                        act=v1[0]-v2[0]-v3[0]+v4[0]-v5[0]-v6[0]
                        bud=v1[1]-v2[1]-v3[1]+v4[1]-v5[1]-v6[1]
                        if act and bud:
                            p=float(act)/float(bud)
                        ps=0
                        psn=0
                        if act > 0 :
                            acts=float(act)*.0275
                            buds=float(bud)*.0275
                            
                            if acts and buds:
                                ps=float(acts)/float(buds)

                            actsn=act-acts
                            budsn=bud-buds
                            
                            if actsn and budsn:
                                psn=float(actsn)/float(budsn)
                        else:
                            acts=0
                            buds=0
                            actsn=act
                            budsn=bud
                        html+='<td class="text-right">'+str(frappe.utils.fmt_money(act))+'</td>'
                        html3+='<td class="text-right">'+str(frappe.utils.fmt_money(acts))+'</td>'
                        html4+='<td class="text-right">'+str(frappe.utils.fmt_money(actsn))+'</td>'

                        html2+='<td class="text-right">'+str(frappe.utils.fmt_money(bud))+'</td>'
                        html32+='<td class="text-right">'+str(frappe.utils.fmt_money(buds))+'</td>'
                        html42+='<td class="text-right">'+str(frappe.utils.fmt_money(budsn))+'</td>'

                    v1=prof.get('total')
                    v2=gprof.get('total')
                    v3=ind.get('total')
                    v4=othinc.get('total')
                    v5=othexp.get('total')
                    v6=d.get('total')
                    act=v1[0]-v2[0]-v3[0]+v4[0]-v5[0]-v6[0]
                    bud=v1[1]-v2[1]-v3[1]+v4[1]-v5[1]-v6[1]
                    if act and bud:
                        p=float(act)/float(bud)
                    

                    ps=0
                    psn=0
                    if act > 0 :
                        acts=float(act)*.0275
                        buds=float(bud)*.0275
                        
                        if acts and buds:
                            ps=float(acts)/float(buds)

                        actsn=act-acts
                        budsn=bud-buds
                        
                        if actsn and budsn:
                            psn=float(actsn)/float(budsn)
                    else:
                        acts=0
                        buds=0
                        actsn=act
                        budsn=bud

                    html+='<td class="text-right">'+str(frappe.utils.fmt_money(act))+'</td>'
                    html3+='<td class="text-right">'+str(frappe.utils.fmt_money(acts))+'</td>'
                    html4+='<td class="text-right">'+str(frappe.utils.fmt_money(actsn))+'</td>'

                    html2+='<td class="text-right">'+str(frappe.utils.fmt_money(bud))+'</td>'
                    html32+='<td class="text-right">'+str(frappe.utils.fmt_money(buds))+'</td>'
                    html42+='<td class="text-right">'+str(frappe.utils.fmt_money(budsn))+'</td>'

                    html+='<td></td>'+html2+'</tr >'
                    html3+='<td></td>'+html32+'</tr >'
                    html4+='<td></td>'+html42+'</tr >'
                    html+=html3+html4
                #----------------------------------------------------------------------------------------
                totals.append(d)

        html+= '</table></div>'    

        #----------------------------------------------------------------------------    
        tbss='budget'+str(tbc)
        repdatahtml.update({tbss:html})
        tbc+=1

        dcon=[]
    
        df = pd.read_html(html,header=0)[0]
        headconsol=headconsol+headconsol2        
        dcon.append(bugtit)
        dcon.append(headconsol)
        dcon=dcon+df.values.tolist()
        for index, val in enumerate(dcon):
            for index2, valu in enumerate(val):
                
                pattern = r"^[-+]?[0-9]*\.?[0-9]+$"
                match = re.match(pattern, str(dcon[index][index2]))
                if match:
                    dcon[index][index2]=float(dcon[index][index2])
        
        nam=str(frappe.session.user)+tbd.get('title')
        consolid=json.dumps(dcon)
        cache = frappe.cache()
        cache.set(nam, consolid) 
    
        
    
    
    return repdatahtml

def get_child_acc(account):
    accc=[]
    #accc.append(account.replace("'", "\\'"))
    account=account.replace("'", "\\'")
    acsql=frappe.db.sql(""" select name,is_group from  `tabAccount` where parent_account='{0}' """.format(account),as_dict=1,debug=0)
    if acsql:
        for ac in acsql:
            if ac.is_group:
                accc=accc+get_child_acc(ac.name)
            else:
                accc.append(ac.name.replace("'", "\\'"))
    return accc

def get_parent_group(gp):
    return frappe.db.get_value('MIS Report Group Labels',{'group_label':gp},'parent_mis_report_group_labels')

@frappe.whitelist()
def get_tab_list(company):
    con=['CONSOLIDATE']
    tbls=frappe.db.get_all('MIS Report Settings',filters={'company':company},fields=['title'],order_by='display_order',pluck='title')
    return con+tbls


@frappe.whitelist()
def down_report(company,fiscal_year,fiscal_month):
    
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl import Workbook    
    from openpyxl.chart import (
        PieChart, LineChart,
        ProjectedPieChart,
        Reference
    )
    
    from openpyxl.styles import Font
    from openpyxl.chart.axis import DateAxis
    #from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.styles import PatternFill
    from openpyxl.styles import Border, Side

    ft = Font(bold=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidate"    
    #ws1 = wb.create_sheet("Budget")
    yellow = "00D5D7D9"
    black="00000000"
    thin = Side(border_style="thin", color=black)
    double = Side(border_style="double", color=black)
    thick = Side(border_style="thick", color=black)
    
            
    nam=str(frappe.session.user)+'CONSOLIDATE'   
    cache = frappe.cache()
    cachval=cache.get(nam) 
    cachval=json.loads(cachval)
    for row in cachval:
       ws.append(row) 

    rrowlen=len(cachval)
    rcollen=len(cachval[3])

    rlbl=getColumnName(rcollen)
    rhd="A2:"+str(rlbl)+str(2)
    
    for row in ws[rhd]:
        for cell in row:
            cell.font = ft
            cell.fill = PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")

    rhd="A1:"+str(rlbl)+str(rrowlen)
    for row in ws[rhd]:
        rowhed=0
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if "Total" in str(cell.value) or "PROFIT/LOSS" in str(cell.value) or "GROSS PROFIT" in str(cell.value) or "AFTER DEPRECIATION PROFIT /LOSS" in str(cell.value): 
                rowhed=1
            if rowhed==1:
                cell.font = ft
                cell.fill = PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")

    rhd="A2:A"+str(rrowlen)
    for row in ws[rhd]:
        for cell in row:
            #cell.font = ft
            cell.fill = PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")
    #-----------------------------------------
    tablist=get_tab_list(company)
    for tbls in tablist:
        if tbls!='CONSOLIDATE':
            nam=str(frappe.session.user)+tbls   
            cache = frappe.cache()
            cachval=cache.get(nam) 
            cachval=json.loads(cachval)
            ws2 = wb.create_sheet(tbls)
            for row in cachval:
                ws2.append(row) 

            rrowlen=len(cachval)
            rcollen=len(cachval[3])

            rlbl=getColumnName(rcollen)
            rhd="A2:"+str(rlbl)+str(2)
            
            for row in ws2[rhd]:
                for cell in row:
                    cell.font = ft
                    cell.fill = PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")

            rhd="A1:"+str(rlbl)+str(rrowlen)
            for row in ws2[rhd]:
                rowhed=0
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    if "Total" in str(cell.value) or "NET PROFIT/(LOSS)" in str(cell.value) or "GROSS PROFIT" in str(cell.value) or "AFTER DEPRECIATION PROFIT /LOSS" in str(cell.value): 
                        rowhed=1
                    if rowhed==1:
                        cell.font = ft
                        cell.fill = PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")

            rhd="A2:A"+str(rrowlen)
            for row in ws2[rhd]:
                for cell in row:
                    #cell.font = ft
                    cell.fill = PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")

    file_name=company+'.xlsx'    
    temp_file=os.path.join(frappe.utils.get_bench_path(), "logs", file_name)
    wb.save(temp_file)
    return temp_file

def getColumnName(n):
 
    # initialize output string as empty
    result = ''
 
    while n > 0:
 
        # find the index of the next letter and concatenate the letter
        # to the solution
 
        # here index 0 corresponds to 'A', and 25 corresponds to 'Z'
        index = (n - 1) % 26
        result += chr(index + ord('A'))
        n = (n - 1) // 26
 
    return result[::-1]

@frappe.whitelist()
def down_file(file=None):
    from frappe.utils.file_manager import download_file
    file_name = os.path.basename(file)
    with open(file, "rb") as fileobj:
        filedata = fileobj.read()
    frappe.response['content_type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    frappe.response['content_disposition'] = 'attachment; filename="{0}"'.format(file_name)
    frappe.local.response.filename = file_name
    frappe.local.response.filecontent = filedata
    frappe.local.response.type = "download"